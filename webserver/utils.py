import gspread
from google.oauth2.service_account import Credentials
import base64
import os
import io
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import json, hashlib, os
from datetime import datetime
import pickle
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

SCOPES = ['https://www.googleapis.com/auth/drive.file']

def get_gsheet_client(json_keyfile):
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(json_keyfile, scopes=scopes)
    client = gspread.authorize(creds)
    return client

def append_order_to_gsheet(sheet_id, worksheet_name, order_data):
    default_key = os.path.join(os.path.dirname(__file__), 'festive-post-464106-p3-d250f2280eae.json')
    key_path = os.environ.get('GOOGLE_SERVICE_ACCOUNT_FILE', default_key)
    client = get_gsheet_client(key_path)
    sh = client.open_by_key(sheet_id)
    ws = sh.worksheet(worksheet_name)
    # Nếu sheet trống (chưa có header), thêm header vào dòng 1
    if ws.row_count == 0 or not ws.get_all_values():
        ws.append_row([
            "id", "Đơn hàng số", "Loại giấy", "Số lượng 1 đơn", "Thời gian", "QR", "trạng thái"
        ])
    ws.append_row([
        order_data.get('id', ''),
        order_data.get('Đơn hàng số', ''),
        order_data.get('loai_giay', ''),
        order_data.get('so_luong', 0),
        order_data.get('thoi_gian_luu', ''),
        order_data.get('qr_codes', ''),  # QR code data
        order_data.get('trạng_thái', '')  # Trạng thái đơn hàng
    ], value_input_option="USER_ENTERED")
def qr_bytes_to_base64(qr_bytes):
    return base64.b64encode(qr_bytes).decode('utf-8')

def get_drive_service():
    creds = None
    if os.path.exists('token_drive.pickle'):
        with open('token_drive.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and hasattr(creds, 'expired') and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(r'D:\Visualstudiocode\QRManaShinYi\webserver\google_oauth_config.json', SCOPES)
            creds = flow.run_local_server(port=8000)
        with open('token_drive.pickle', 'wb') as token:
            pickle.dump(creds, token)
    service = build('drive', 'v3', credentials=creds)
    return service


def get_drive_service_service_account():
    """Return a Google Drive service built from a service account JSON file.
    Looks for env var GOOGLE_SERVICE_ACCOUNT_FILE or falls back to the bundled key.
    """
    try:
        default_key = os.path.join(os.path.dirname(__file__), 'festive-post-464106-p3-d250f2280eae.json')
        key_path = os.environ.get('GOOGLE_SERVICE_ACCOUNT_FILE', default_key)
        # Use full Drive scope to allow permission changes
        scopes = ['https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(key_path, scopes=scopes)
        service = build('drive', 'v3', credentials=creds)
        return service
    except Exception as e:
        print('Could not create Drive service from service account:', e)
        return None

def upload_file_to_drive(local_file_path, drive_filename=None):
    # Prefer service-account-based Drive service if a key is available.
    service = None
    try:
        sa_key = os.environ.get('GOOGLE_SERVICE_ACCOUNT_FILE')
        if sa_key and os.path.exists(sa_key):
            service = get_drive_service_service_account()
    except Exception:
        service = None
    if not service:
        # fallback to user-installed-app OAuth flow (token_drive.pickle)
        service = get_drive_service()
    file_metadata = {'name': drive_filename or os.path.basename(local_file_path)}
    media = MediaFileUpload(local_file_path, resumable=True)
    file = service.files().create(body=file_metadata, media_body=media, fields='id,webViewLink,webContentLink').execute()
    file_id = file.get('id')
    webview = file.get('webViewLink') or file.get('webContentLink')
    # Try to make the file accessible to anyone with the link. This works for
    # both service account and regular OAuth credentials when allowed by Drive.
    try:
        service.permissions().create(fileId=file_id, body={"role": "reader", "type": "anyone"}).execute()
    except Exception as e:
        print('Could not set file permission to anyone:', e)
    # Ensure we can return a usable link
    try:
        f2 = service.files().get(fileId=file_id, fields='id,webViewLink,webContentLink').execute()
        webview = f2.get('webViewLink') or f2.get('webContentLink') or webview
    except Exception:
        pass
    # Construct a reliable direct link if Drive didn't return webViewLink
    if not webview and file_id:
        webview = f'https://drive.google.com/uc?export=view&id={file_id}'
    print(f"Uploaded file to Google Drive: {webview}")
    return webview

# This file handles user management for the webserver, including loading, saving users,
USERS_FILE = os.path.join(os.path.dirname(__file__), "users.json")

def load_users():
    if not os.path.exists(USERS_FILE):
        return []
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_users(users):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def add_user(username, password, role):
    users = load_users()
    if any(u["username"] == username for u in users):
        return False
    users.append({
        "username": username,
        "password_hash": hash_password(password),
        "role": role
    })
    save_users(users)
    return True

def check_user(username, password):
    users = load_users()
    for u in users:
        if u["username"] == username and u["password_hash"] == hash_password(password):
            return u["role"]
    return None


def generate_qr_image_bytes(data: str):
    """Generate QR code image bytes (PNG) from a string."""
    try:
        import qrcode
        import io
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        bio = io.BytesIO()
        img.save(bio, format='PNG')
        bio.seek(0)
        return bio.read()
    except Exception as e:
        print('QR generation error:', e)
        return None


def append_product_to_gsheet(sheet_id, worksheet_name, product_data):
    """Append a product row to Google Sheet. product_data is a dict with keys:
    ['product_id','name','description','cert_drive_link','qr_drive_link','created_at']
    """
    try:
        default_key = os.path.join(os.path.dirname(__file__), 'festive-post-464106-p3-d250f2280eae.json')
        key_path = os.environ.get('GOOGLE_SERVICE_ACCOUNT_FILE', default_key)
        client = get_gsheet_client(key_path)
        sh = client.open_by_key(sheet_id)
        try:
            ws = sh.worksheet(worksheet_name)
        except Exception:
            # try to create worksheet
            try:
                ws = sh.add_worksheet(title=worksheet_name, rows=1000, cols=20)
            except Exception as e:
                print('Could not get or create worksheet:', e)
                return False
        # Ensure header exists
        if ws.row_count == 0 or not ws.get_all_values():
            ws.append_row(['product_id','name','description','staff_name','staff_id','product_code','issue_date','serial_no','certificate_no','organization','check_count','cert_drive_link','external_link','qr_drive_link','created_at'])
        ws.append_row([
            product_data.get('product_id',''),
            product_data.get('name',''),
            product_data.get('description',''),
            product_data.get('staff_name',''),
            product_data.get('staff_id',''),
            product_data.get('cert_drive_link',''),
            product_data.get('external_link',''),
            product_data.get('qr_drive_link',''),
            product_data.get('created_at','')
        ], value_input_option='USER_ENTERED')
        return True
    except Exception as e:
        print('Error appending product to gsheet:', e)
        return False


def ensure_qr_dir():
    qr_dir = os.path.join(os.path.dirname(__file__), 'static', 'QRmana')
    if not os.path.exists(qr_dir):
        try:
            os.makedirs(qr_dir, exist_ok=True)
        except Exception as e:
            print('Could not create QR directory:', e)
            return None
    return qr_dir


def save_qr_to_local(qr_bytes, filename=None):
    """Save QR bytes to webserver/static/QRmana and return relative url path."""
    qr_dir = ensure_qr_dir()
    if not qr_dir:
        return None
    if not filename:
        filename = f"qr_{int(datetime.now().timestamp())}.png"
    path = os.path.join(qr_dir, filename)
    try:
        with open(path, 'wb') as f:
            f.write(qr_bytes)
        # return a URL path relative to static mount
        return f"/static/QRmana/{filename}"
    except Exception as e:
        print('Error saving QR locally:', e)
        return None


def append_qr_record_to_gsheet(sheet_id, worksheet_name, qr_data):
    """Append a minimal QR record: stt, staff_name, qr_code, note"""
    try:
        default_key = os.path.join(os.path.dirname(__file__), 'festive-post-464106-p3-d250f2280eae.json')
        key_path = os.environ.get('GOOGLE_SERVICE_ACCOUNT_FILE', default_key)
        client = get_gsheet_client(key_path)
        sh = client.open_by_key(sheet_id)
        try:
            ws = sh.worksheet(worksheet_name)
        except Exception:
            try:
                ws = sh.add_worksheet(title=worksheet_name, rows=1000, cols=10)
            except Exception as e:
                print('Could not get/create QR sheet:', e)
                return False
        rows = ws.get_all_values() or []
        if not rows:
            # add header
            ws.append_row(['stt', 'ten_nhan_vien', 'ma_qr', 'ghi_chu'])
            next_stt = 1
        else:
            # compute next stt as number of rows (header included)
            next_stt = len(rows)
            # If first row contains header label 'stt', keep next_stt as len(rows)
        # ensure QR/code fields are strings (avoid URL or other objects)
        qr_code_val = qr_data.get('qr_code', '')
        try:
            qr_code_val = str(qr_code_val)
        except Exception:
            qr_code_val = ''
        ws.append_row([
            str(next_stt),
            qr_data.get('staff_name', ''),
            qr_code_val,
            qr_data.get('note', '')
        ], value_input_option='USER_ENTERED')
        return True
    except Exception as e:
        print('Error appending QR record to gsheet:', e)
        return False


def append_summary_to_gsheet(sheet_id, worksheet_name, summary):
    """Append a concise summary row to Google Sheet with columns:
    STT, ID, Tên sản phẩm, Seri tem, mã qr, thời gian
    summary is a dict with keys: id, name, serial_no, qr_code, time
    """
    try:
        default_key = os.path.join(os.path.dirname(__file__), 'festive-post-464106-p3-d250f2280eae.json')
        key_path = os.environ.get('GOOGLE_SERVICE_ACCOUNT_FILE', default_key)
        client = get_gsheet_client(key_path)
        sh = client.open_by_key(sheet_id)
        try:
            ws = sh.worksheet(worksheet_name)
        except Exception:
            try:
                ws = sh.add_worksheet(title=worksheet_name, rows=1000, cols=10)
            except Exception as e:
                print('Could not create worksheet:', e)
                return False
        rows = ws.get_all_values() or []
        # New header with requested columns: ID QR, Tên sản phẩm, Mã đơn hàng, Seri tem, Ngày nhập, Ngày hồ sơ, Tên nhân viên, Mã nhân viên, Link
        header = ['STT', 'ID QR', 'Tên sản phẩm', 'Mã đơn hàng', 'Seri tem', 'Ngày nhập', 'Ngày hồ sơ', 'Tên nhân viên', 'Mã nhân viên', 'Link']
        if not rows:
            ws.append_row(header)
            next_stt = 1
        else:
            next_stt = len(rows)
        # Prepare values
        created_at = summary.get('created_at', '')
        issue_date = summary.get('issue_date', '')
        # For Google Sheets, write a HYPERLINK formula so the display text is 'link'
        # ensure detail_url is a plain string (request.url_for may return a URL object)
        detail_url = summary.get('detail_url', '')
        if detail_url is None:
            detail_url = ''
        else:
            detail_url = str(detail_url)
        escaped = detail_url.replace('"', '""') if detail_url else ''
        hyperlink_cell = f'=HYPERLINK("{escaped}","link")' if detail_url else ''
        ws.append_row([
            str(next_stt),
            summary.get('id', ''),
            summary.get('name', ''),
            summary.get('product_code', ''),
            summary.get('serial_no', ''),
            created_at,
            issue_date,
            summary.get('staff_name', ''),
            summary.get('staff_id', ''),
            hyperlink_cell
        ], value_input_option='USER_ENTERED')
        return True
    except Exception as e:
        print('Error appending summary to gsheet:', e)
        return False


def append_summary_to_excel(excel_path, summary):
    """Append concise summary row to local Excel file with columns: STT, ID, Tên sản phẩm, Seri tem, Mã QR, Thời gian"""
    try:
        from openpyxl import Workbook, load_workbook
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = 'ProductsSummary'
            # Header: STT, ID QR, Tên sản phẩm, Mã đơn hàng, Seri tem, Ngày nhập, Ngày hồ sơ, Tên nhân viên, Mã nhân viên, Link
            ws.append(['STT', 'ID QR', 'Tên sản phẩm', 'Mã đơn hàng', 'Seri tem', 'Ngày nhập', 'Ngày hồ sơ', 'Tên nhân viên', 'Mã nhân viên', 'Link'])
            wb.save(excel_path)
        wb = load_workbook(excel_path)
        if 'ProductsSummary' not in wb.sheetnames:
            wb.create_sheet('ProductsSummary')
        ws = wb['ProductsSummary']
        next_stt = ws.max_row
        created_at = summary.get('created_at', '')
        issue_date = summary.get('issue_date', '')
        # Append row with placeholder for link; we'll set the hyperlink on the cell after append
        ws.append([
            str(next_stt),
            summary.get('id', ''),
            summary.get('name', ''),
            summary.get('product_code', ''),
            summary.get('serial_no', ''),
            created_at,
            issue_date,
            summary.get('staff_name', ''),
            summary.get('staff_id', ''),
            ''
        ])
        # set hyperlink on the last cell of the appended row
        try:
            from openpyxl.styles import Font
            row_idx = ws.max_row
            # Link is now the 10th column (STT=1, ID=2, Name=3, product_code=4, serial_no=5, created_at=6, issue_date=7, staff_name=8, staff_id=9, Link=10)
            col_idx = 10
            url = summary.get('detail_url', '')
            if url:
                cell = ws.cell(row=row_idx, column=col_idx)
                # store the full URL as plain text so the user can copy/paste it into a browser
                cell.value = str(url)
                # ensure the cell is formatted as text and does not have an openpyxl hyperlink object
                try:
                    cell.hyperlink = None
                except Exception:
                    pass
                try:
                    cell.number_format = '@'
                except Exception:
                    pass
        except Exception:
            # If openpyxl styling/hyperlink fails, ignore and leave empty
            pass
        try:
            wb.save(excel_path)
            return True
        except PermissionError as pe:
            # File may be locked (e.g., opened in Excel) or not writable. Try a fallback filename.
            try:
                alt_name = f"products_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                alt_path = os.path.join(os.path.dirname(excel_path), alt_name)
                wb.save(alt_path)
                print(f"Permission denied saving to {excel_path}. Saved to fallback file: {alt_path}")
                return True
            except Exception as e:
                print('Error saving excel summary (fallback failed):', e)
                return False
    except Exception as e:
        print('Error appending summary to excel:', e)
        return False


def append_product_to_excel(excel_path, product_data):
    """Append a product row to a local Excel file (creates file/sheet if not exists)."""
    try:
        from openpyxl import Workbook, load_workbook
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = 'Products'
            ws.append(['product_id','name','description','staff_name','staff_id','cert_drive_link','external_link','qr_drive_link','created_at'])
            wb.save(excel_path)
        wb = load_workbook(excel_path)
        if 'Products' not in wb.sheetnames:
            wb.create_sheet('Products')
        ws = wb['Products']
        ws.append([
            product_data.get('product_id',''),
            product_data.get('name',''),
            product_data.get('description',''),
            product_data.get('staff_name',''),
            product_data.get('staff_id',''),
            product_data.get('cert_drive_link',''),
            product_data.get('external_link',''),
            product_data.get('qr_drive_link',''),
            product_data.get('created_at','')
        ])
        wb.save(excel_path)
        return True
    except Exception as e:
        print('Error appending product to excel:', e)
        return False